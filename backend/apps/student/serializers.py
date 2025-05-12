from rest_framework import serializers
from apps.student.models import Student, Intake
from apps.tracks.models import Track
import openpyxl
from django.core.mail import send_mail
from django.conf import settings
import secrets
import string
from django.core.exceptions import ValidationError
import logging
from django.db import transaction
from django.core.validators import validate_email
from io import BytesIO

logger = logging.getLogger(__name__)

class IntakeSerializer(serializers.ModelSerializer):
    track = serializers.PrimaryKeyRelatedField(
        queryset=Track.objects.all(),
        required=True,
        allow_null=False
    )

    class Meta:
        model = Intake
        fields = ['id', 'name', 'track']

    def validate(self, data):
        name = data.get('name')
        track = data.get('track')
        if not track:
            raise serializers.ValidationError({"track": "Track is required."})
        if Intake.objects.filter(name=name, track=track).exists():
            raise serializers.ValidationError(
                f"Intake '{name}' already exists for track '{track.name}'."
            )
        return data

class ExcelUploadSerializer(serializers.Serializer):
    excel_file = serializers.FileField()
    track_id = serializers.PrimaryKeyRelatedField(
        queryset=Track.objects.all(),
        required=True,
        source='track'
    )
    intake_name = serializers.CharField(required=True, max_length=100)

    def validate_excel_file(self, value):
        if not value.name.lower().endswith(('.xlsx', '.xls')):
            raise ValidationError("Only .xlsx or .xls files allowed")
        try:
            content = value.read()
            value.seek(0)
            wb = openpyxl.load_workbook(filename=BytesIO(content))
            sheet = wb.active
        except Exception as e:
            logger.error(f"Excel read error: {str(e)}")
            raise ValidationError(f"Invalid Excel file: {str(e)}")
        try:
            header_row = [str(cell.value).strip().lower() for cell in sheet[1]]
            required_columns = ['first name', 'last name', 'email', 'role']
            missing = [col for col in required_columns if col not in header_row]
            if missing:
                raise ValidationError(
                    f"Missing columns: {', '.join(missing).title()}. "
                    f"Found columns: {', '.join(header_row)}"
                )
        except IndexError:
            raise ValidationError("Excel file has no header row")
        if sheet.max_row < 2:
            raise ValidationError("Excel file contains no data rows")
        return value

    def validate_intake_name(self, value):
        return value.strip()

    def validate(self, data):
        track = data.get('track')
        intake_name = data.get('intake_name')
        if not track:
            raise ValidationError({"track": "Track is required"})
        return data

    def _process_excel_file(self):
        excel_file = self.validated_data['excel_file']
        track = self.validated_data['track']
        intake_name = self.validated_data['intake_name']

        intake, created = Intake.objects.get_or_create(
            name=intake_name,
            track=track,
            defaults={'name': intake_name, 'track': track}
        )

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            required_columns = ['First Name', 'Last Name', 'Email', 'Role']
            if not all(col in header_row for col in required_columns):
                raise ValidationError(
                    f"Excel file missing required columns. Found: {header_row}, Required: {required_columns}"
                )
        except Exception as e:
            e.sheet = getattr(sheet, 'title', 'Unknown')
            logger.error(f"Failed to open Excel file: {str(e)}")
            raise ValidationError("Invalid Excel file format or structure")

        students_to_create = []
        existing_emails_in_intake = set(
            Student.objects.filter(intake=intake).values_list('email', flat=True)
        )
        created_students = []
        email_password_map = {}
        errors = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row[:4]):
                continue
            try:
                if len(row) < 4:
                    errors.append(f"Row {row_num}: Insufficient data (expected 4 columns, got {len(row)})")
                    continue
                first_name, last_name, email, role = row[:4]
                email = (email or "").strip().lower()
                if not first_name:
                    errors.append(f"Row {row_num}: Missing first name")
                    continue
                if not last_name:
                    errors.append(f"Row {row_num}: Missing last name")
                    continue
                if not email:
                    errors.append(f"Row {row_num}: Missing email")
                    continue
                if not self._validate_email(email):
                    errors.append(f"Row {row_num}: Invalid email format '{email}'")
                    continue
                if email in existing_emails_in_intake:
                    errors.append(f"Row {row_num}: Email '{email}' already exists in intake '{intake_name}'")
                    continue
                password = self._generate_password()
                verification_code = self._generate_verification_code()
                student = Student(
                    username=self._generate_username(email),
                    first_name=str(first_name).strip(),
                    last_name=str(last_name).strip(),
                    email=email,
                    intake=intake,
                    role=(str(role) if role else 'student').strip().lower(),
                    track=track,
                    verification_code=verification_code,
                    verified=False
                )
                student.set_password(password)
                students_to_create.append(student)
                existing_emails_in_intake.add(email)
                email_password_map[email] = password
            except Exception as e:
                errors.append(f"Row {row_num}: Error processing - {str(e)}")
                logger.error(f"Row {row_num}: Error processing - {str(e)}")
                continue

        if errors and not students_to_create:
            raise ValidationError({
                'detail': 'All rows failed validation',
                'errors': errors
            })

        try:
            with transaction.atomic():
                created_students = Student.objects.bulk_create(students_to_create)
                self._send_verification_emails(created_students, email_password_map)
        except Exception as e:
            logger.error(f"Database error during bulk create: {str(e)}")
            raise ValidationError(f"Failed to create student records: {str(e)}")

        return {
            "status": "partial_success" if errors else "success",
            "created_count": len(created_students),
            "error_count": len(errors),
            "students": StudentSerializer(created_students, many=True).data,
            "errors": errors if errors else []
        }

    def _validate_email(self, email):
        try:
            validate_email(email)
            return True
        except ValidationError:
            return False

    def _generate_password(self):
        alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
        return ''.join(secrets.choice(alphabet) for _ in range(12))

    def _generate_verification_code(self):
        return secrets.token_urlsafe(24)

    def _generate_username(self, email):
        return email.split('@')[0]

    def _send_verification_emails(self, students, email_password_map):
        for student in students:
            try:
                password = email_password_map.get(student.email)
                if password:
                    verification_url = f"{settings.SITE_URL}/api/student/verify/{student.verification_code}/"
                    subject = "Verify Your Student Account"
                    message = f"""
                    Hello {student.first_name},

                    Your student account has been created. Please verify your email by visiting:
                    {verification_url}

                    Your login information:
                    - Email: {student.email}
                    - Temporary Password: {password}
                    - Intake ID: {student.intake.id}
                    - Track: {student.track.name}

                    Use these credentials to log in as a student. After verification, you can change your password.
                    """
                    send_mail(
                        subject,
                        message.strip(),
                        settings.DEFAULT_FROM_EMAIL,
                        [student.email],
                        fail_silently=False
                    )
                    logger.info(f"Verification email sent to {student.email}")
            except Exception as e:
                logger.error(f"Failed to send email to {student.email}: {str(e)}")

    def save(self, **kwargs):
        try:
            return self._process_excel_file()
        except ValidationError as e:
            raise
        except Exception as e:
            logger.exception("Unexpected error during Excel processing")
            raise ValidationError("Failed to process Excel file")

class StudentSerializer(serializers.ModelSerializer):
    intake2 = IntakeSerializer(read_only=True)
    
    track = serializers.StringRelatedField(read_only=True)
    track_id = serializers.PrimaryKeyRelatedField(
        queryset=Track.objects.all(),
        write_only=True,
        required=True,
        source='track'
    )
    intake = serializers.StringRelatedField(read_only=True)
    intake_id = serializers.PrimaryKeyRelatedField(
        queryset=Intake.objects.all(),
        write_only=True,
        required=True,
        source='intake'
    )
    password = serializers.CharField(
        write_only=True,
        required=False,
        allow_blank=True,
        style={'input_type': 'password'}
    )

    class Meta:
        model = Student
        fields = [
            'id', 'username', 'email', 'first_name', 'last_name',
            'role', 'track', 'track_id', 'intake', 'intake_id', 
            'password', 'is_active', 'verified', 'date_joined', 'intake2'
        ]
        read_only_fields = [
            'id', 'is_active', 'verified', 'date_joined'
        ]
        extra_kwargs = {
            'email': {'required': True},
            'username': {'required': False, 'allow_blank': True},
            'first_name': {'required': True},
            'last_name': {'required': True}
        }

    def validate_email(self, value):
        logger.debug(f"Validating email: {value}")
        intake_id = self.context['request'].data.get('intake_id') if self.context.get('request') else self.initial_data.get('intake_id')
        if intake_id:
            try:
                Intake.objects.get(id=intake_id)
            except Intake.DoesNotExist:
                raise serializers.ValidationError("Invalid intake ID")
        if self.instance and self.instance.email == value and (not intake_id or self.instance.intake_id == int(intake_id)):
            return value
        if intake_id and Student.objects.filter(email=value, intake_id=intake_id).exists():
            raise serializers.ValidationError(f"Email already exists in this intake")
        try:
            validate_email(value)
        except ValidationError:
            raise serializers.ValidationError("Enter a valid email address")
        return value.lower().strip()

    def validate(self, data):
        logger.debug(f"Serializer data: {data}")
        if not data.get('username'):
            email = data.get('email', '')
            data['username'] = email.split('@')[0] if email else ''
        track = data.get('track')
        intake = data.get('intake')
        if track and intake and intake.track != track:
            raise serializers.ValidationError("Intake must belong to the selected track")
        return data

    def create(self, validated_data):
        logger.debug(f"Creating student with validated data: {validated_data}")
        password = validated_data.pop('password', None)
        track = validated_data.pop('track')
        intake = validated_data.pop('intake')
        student = Student(**validated_data)
        student.track = track
        student.intake = intake
        if password:
            student.set_password(password)
        else:
            temp_password = self._generate_temp_password()
            student.set_password(temp_password)
        student.verification_code = self._generate_verification_code()
        student.save()
        self._send_verification_email(student, password or temp_password)
        return student

    def update(self, instance, validated_data):
        logger.debug(f"Updating student with validated data: {validated_data}")
        password = validated_data.pop('password', None)
        track = validated_data.pop('track', None)
        intake = validated_data.pop('intake', None)
        for attr, value in validated_data.items():
            setattr(instance, attr, value)
        if track is not None:
            instance.track = track
        if intake is not None:
            instance.intake = intake
        if password:
            instance.set_password(password)
        instance.save()
        return instance

    def _generate_temp_password(self):
        alphabet = string.ascii_letters + string.digits
        return ''.join(secrets.choice(alphabet) for _ in range(10))

    def _generate_verification_code(self):
        return secrets.token_urlsafe(24)

    def _send_verification_email(self, student, password):
        try:
            verification_url = f"{settings.SITE_URL}/api/student/verify/{student.verification_code}/"
            subject = "Verify Your Student Account"
            message = f"""
            Hello {student.first_name},

            Your student account has been created. Please verify your email by visiting:
            {verification_url}

            Your login information:
            - Email: {student.email}
            - Temporary Password: {password}
            - Intake ID: {student.intake.id}
            - Track: {student.track.name}

            Use these credentials to log in as a student. After verification, you can change your password.
            """
            send_mail(
                subject,
                message.strip(),
                settings.DEFAULT_FROM_EMAIL,
                [student.email],
                fail_silently=False
            )
            logger.info(f"Verification email sent to {student.email}")
        except Exception as e:
            logger.error(f"Failed to send verification email to {student.email}: {str(e)}")

class DashboardSerializer(serializers.ModelSerializer):
    upcoming_assignments = serializers.SerializerMethodField()
    courses = serializers.SerializerMethodField()
    
    class Meta:
        model = Student
        fields = ['id', 'username', 'email', 'first_name', 'last_name', 
                'upcoming_assignments', 'courses', 'intake']
    
    def get_upcoming_assignments(self, obj):
        return []
    
    def get_courses(self, obj):
        return []
    
class MinimalStudentSerializer(serializers.ModelSerializer):
    intake = serializers.StringRelatedField()
    
    class Meta:
        model = Student
        fields = ['id', 'full_name', 'email', 'intake']
        
class StudentSubmissionStatusSerializer(serializers.ModelSerializer):
    intake = serializers.StringRelatedField()
    has_submitted = serializers.SerializerMethodField()
    submission_details = serializers.SerializerMethodField()
    
    class Meta:
        model = Student
        fields = ['id', 'full_name', 'email', 'intake', 'has_submitted', 'submission_details']
    
    def get_has_submitted(self, student):
        assignment_id = self.context.get('assignment_id')
        return AssignmentStudent.objects.filter(
            student=student,
            assignment_id=assignment_id,
            submitted=True
        ).exists()
    
    def get_submission_details(self, student):
        assignment_id = self.context.get('assignment_id')
        try:
            assignment_student = AssignmentStudent.objects.get(
                student=student,
                assignment_id=assignment_id
            )
            if assignment_student.submitted:
                submission = Submission.objects.filter(
                    assignment_student=assignment_student
                ).first()
                return {
                    'submitted_at': assignment_student.submission_date,
                    'file_url': submission.file.url if submission else None
                }
        except AssignmentStudent.DoesNotExist:
            return None
        return None